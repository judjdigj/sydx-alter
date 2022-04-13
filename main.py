import openpyxl

def song_find(key):
    status = 0
    row_count = 0
    for row in sheet.iter_rows():
        row_count = row_count + 1
        if status == 0:
            for cell in row:
                if cell.value == key:
                    print(sheet.cell(row_count, 1).value)
                    status = 1
                    break
        else:
            break

book = openpyxl.load_workbook('song_alter_test.xlsx')
sheet = book.active
song_name = 0
song_alter = 0
row_count = 0
col_count = 0

print("Mode Select")
print("1 for add alter name")
print("2 for check alter name")
mode_select = input()

if mode_select == "2":
    print("key")
    song_find(input())
