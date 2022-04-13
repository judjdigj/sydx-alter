import openpyxl

book = openpyxl.load_workbook('song_alter_test.xlsx') #Form contain song's original and alternative name. There's only one in the project.
sheet = book.active

def song_find(key):
    #You only need one arguement, which is the keyword of the song's alternative/original name 
    row_count = 0
    O = 0
    for row in sheet.iter_rows():
        row_count = row_count + 1
        for cell in row:
            if cell.value == key:
                return sheet.cell(row_count, 1).value   #Retuen the original name, which in the first column.
            elif cell.value == None:
                break

    return "no found" #You can change to whatever you want for the bot.