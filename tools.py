import openpyxl
from openpyxl.comments import Comment

book = openpyxl.load_workbook('song_alter_test.xlsx') #Form contain song's original and alternative name. There's only one in the project.
sheet = book.active

def song_find(key):
    #You only need one arguement, which is the keyword of the song's alternative/original name 
    row_count = 0
    for row in sheet.iter_rows():
        row_count = row_count + 1
        for cell in row:
            if cell.value == key:
                return sheet.cell(row_count, 1).value   #Retuen the original name, which in the first column.
            elif cell.value == None:
                break
    return key #When song no found, return the original search key for the original search engine.

def alter_add(new, exist, qqid):
    #You need 3 arguement, new name, exisied name, and the uploader's QQID
    exist = song_find(exist)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == exist:
                for cell in row:
                    if sheet.cell(cell.row, cell.column+1).value == None:
                        sheet.cell(cell.row, cell.column+1).value = new
                        comment = Comment(qqid, 'Ricebot')
                        comment.width = 100
                        comment.height = 10
                        sheet.cell(cell.row, cell.column+1).comment = comment #Add uploader's QQID as comment
                        book.save('song_alter_test.xlsx')
                        return 'added'     #add alternative name
                    elif sheet.cell(cell.row, cell.column+1).value == new:
                        return "already in library"   #Check for existing alternative name
            elif cell.value == None:
                break
    return 'song no found'   #when no song found